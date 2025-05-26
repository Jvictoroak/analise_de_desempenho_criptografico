import time
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import padding as asymmetric_padding
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding as symmetric_padding
from cryptography.hazmat.backends import default_backend
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# Texto a ser cifrado
texto_original = "RSA eh um algoritmo que leva o nome de 3 professores do MIT: Rivest, Shamir e Adleman"
texto_bytes = texto_original.encode('utf-8')

# Função para medir tempo de execução
def medir_tempo(funcao, *args):
    tempos = []
    for _ in range(3):
        inicio = time.time()
        funcao(*args)
        fim = time.time()
        tempos.append(fim - inicio)
    return tempos

# Função para dividir o texto em blocos adequados para RSA
def dividir_em_blocos(dados, tamanho_bloco):
    return [dados[i:i+tamanho_bloco] for i in range(0, len(dados), tamanho_bloco)]

# RSA
def testar_rsa(tamanho_chave):
    def funcao():
        # Gerar chaves
        private_key = rsa.generate_private_key(
            public_exponent=65537,
            key_size=tamanho_chave,
            backend=default_backend()
        )
        public_key = private_key.public_key()
        
        max_tamanho = (tamanho_chave // 8) - 66
        blocos = dividir_em_blocos(texto_bytes, max_tamanho) if len(texto_bytes) > max_tamanho else [texto_bytes]
        
        ciphertexts = []
        for bloco in blocos:
            ciphertext = public_key.encrypt(
                bloco,
                asymmetric_padding.OAEP(
                    mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
            ciphertexts.append(ciphertext)
        
        for ciphertext in ciphertexts:
            private_key.decrypt(
                ciphertext,
                asymmetric_padding.OAEP(
                    mgf=asymmetric_padding.MGF1(algorithm=hashes.SHA256()),
                    algorithm=hashes.SHA256(),
                    label=None
                )
            )
    return medir_tempo(funcao)

# AES
def testar_aes(tamanho_chave):
    def funcao():
        key = os.urandom(tamanho_chave // 8)
        iv = os.urandom(16)
        
        padder = symmetric_padding.PKCS7(128).padder()
        padded_data = padder.update(texto_bytes) + padder.finalize()
        
        cipher = Cipher(
            algorithms.AES(key),
            modes.CBC(iv),
            backend=default_backend()
        )
        encryptor = cipher.encryptor()
        ciphertext = encryptor.update(padded_data) + encryptor.finalize()
        
        decryptor = cipher.decryptor()
        decrypted_data = decryptor.update(ciphertext) + decryptor.finalize()
        unpadder = symmetric_padding.PKCS7(128).unpadder()
        unpadder.update(decrypted_data) + unpadder.finalize()
    
    return medir_tempo(funcao)

# Executar testes
print("Iniciando testes...")
resultados = {
    "RSA 1024": testar_rsa(1024),
    "RSA 2048": testar_rsa(2048),
    "RSA 4096": testar_rsa(4096),
    "RSA 8192": testar_rsa(8192),
    "AES 128": testar_aes(128),
    "AES 256": testar_aes(256),
}

# Criar arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Resultados Criptografia"

# Cabeçalhos
cabecalhos = ['Algoritmo', 'Execução 1 (s)', 'Execução 2 (s)', 'Execução 3 (s)', 'Média (s)']
for col, cabecalho in enumerate(cabecalhos, 1):
    ws.cell(row=1, column=col, value=cabecalho).font = Font(bold=True)

# Dados
for row, (algoritmo, tempos) in enumerate(resultados.items(), 2):
    media = sum(tempos) / len(tempos)
    ws.cell(row=row, column=1, value=algoritmo)
    for col, tempo in enumerate(tempos, 2):
        ws.cell(row=row, column=col, value=tempo)
    ws.cell(row=row, column=5, value=media)

# Formatar números
for row in ws.iter_rows(min_row=2, max_col=5, max_row=len(resultados)+1):
    for cell in row[1:]:  # A partir da segunda coluna
        cell.number_format = '0.000000'

# Ajustar largura das colunas
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

# Salvar arquivo
nome_arquivo = "resultados_criptografia.xlsx"
wb.save(nome_arquivo)
print(f"\nPlanilha Excel '{nome_arquivo}' gerada com sucesso.")

# Exibir resultados no console
print("\nResultados dos tempos de execução (segundos):")
for algoritmo, tempos in resultados.items():
    media = sum(tempos) / len(tempos)
    print(f"{algoritmo}:")
    print(f"  Execuções: {[f'{t:.6f}' for t in tempos]}")
    print(f"  Média: {media:.6f}s")